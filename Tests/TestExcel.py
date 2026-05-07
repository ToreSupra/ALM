from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="Employee Report")

ws.cell(row=3, column=1, value="Name")
ws.cell(row=3, column=2, value="Alice")

ws.cell(row=4, column=1, value="Age")
ws.cell(row=4, column=2, value=28)

ws.cell(row=5, column=1, value="Department")
ws.cell(row=5, column=2, value="Engineering")

wb.save("cell_by_cell.xlsx")