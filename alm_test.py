"""
Laser Measurement State Machine
================================
States:
  INIT → CHECK_DEVICE_1 → CHECK_DEVICE_2 → LASER_CHECK
  → COARSE_SCAN → FIND_MAX → FINE_SCAN → PLOT_POWER_ANGLE
  → GOTO_MAX → PRECISION_SCAN → DWELL → PLOT_POWER_TIME
  → OUTPUT_FILE → DONE
  (LASER_CHECK → QUIT on 'no')

Device #1 : motor controller  (stub: Device1)
Device #2 : power meter       (stub: Device2)
"""

from __future__ import annotations

from openpyxl import Workbook # For Excel files
import time
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Optional
from typing import Any
import argparse
import json
import sys

import numpy as np

from pymeasure.instruments.thorlabs import ThorlabsPM100USB
from pymeasure.instruments.yokogawa.aq6370series import AQ6375B

# ---------------------------------------------------------------------------
# State enum
# ---------------------------------------------------------------------------

class State(Enum):
    INIT            = auto()
    LOAD_SETTINGS   = auto()
    CHECK_OSA       = auto()
    CHECK_POWERMETER= auto()
    SKYLINE_CHECK   = auto()
    WARMUP_LASER    = auto()
    RAMPUP_LASER    = auto()
    WAIT            = auto()
    MEASURE         = auto()
    OUTPUT_FILE     = auto()
    DONE            = auto()
    QUIT            = auto()
    ERROR           = auto()


# ---------------------------------------------------------------------------
# Measurement context  (shared data between states)
# ---------------------------------------------------------------------------

@dataclass
class Context:
    pm: ThorlabsPM100USB = field(default_factory=ThorlabsPM100USB)
    osa: AQ6375B = field(default_factory=AQ6375B)

    # coarse scan
    coarse_angles:  list = field(default_factory=list)
    coarse_powers:  list = field(default_factory=list)

    # fine scan (power vs angle)
    fine_angles:    list = field(default_factory=list)
    fine_powers:    list = field(default_factory=list)

    # precision scan (90 deg)
    prec_angles:    list = field(default_factory=list)
    prec_powers:    list = field(default_factory=list)

    # dwell (power vs time)
    dwell_times:    list = field(default_factory=list)
    dwell_powers:   list = field(default_factory=list)

    max_position:   float = 0.0     # absolute angle of power maximum
    min_position:   float = 0.0     # absolute angle of power minimum
    current_angle:  float = 0.0     # running absolute position

    output_path:    Optional[Path] = None

    wb:             Workbook = field(default_factory=Workbook)
    ws:             Any = field(init=False)
    ws_row:         int = 1 # next row available

    def __post_init__(self):
        self.ws = self.wb.active

    def resolve_output_path(self) -> Path:
        if self.output_path is not None:
            return self.output_path
        
        i = 0
        while True:
            candidate = Path(f"alm_output{i:04d}.xlsx")
            if not candidate.exists():
                return candidate
            i += 1
    
    def load_settings(self, file_path: Path) -> None:
        """Reads, parses, and validates the JSON settings file into class members."""
        file_path = Path(file_path)

        if not file_path.exists():
            print(f"Error: Settings file '{file_path}' was not found.")
            sys.exit(1)

        if not file_path.suffix == ".json":
            print(f"Error: File '{file_path}' is not a JSON file.")
            sys.exit(1)

        try:
            with open(file_path, 'r') as f:
                settings = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Error: '{file_path}' is not valid JSON. Details: {e.msg} at line {e.lineno}.")
            sys.exit(1)

        self._validate_and_assign(settings)


    def _validate_and_assign(self, settings: dict) -> None:
        """Validates and assigns settings to class members."""

        # --- Required fields ---
        required = ["osa", "osa_sweep_start_wl", "osa_sweep_stop_wl", "osa_res"]
        for key in required:
            if key not in settings:
                print(f"Error: Missing required setting '{key}'.")
                sys.exit(1)

        # --- Validate and assign each field ---
        wavelength = settings["wavelength"]
        if not isinstance(wavelength, (int, float)) or not (200 <= wavelength <= 2000):
            print(f"Error: 'wavelength' must be a number between 200 and 2000, got '{wavelength}'.")
            sys.exit(1)
        self.wavelength = float(wavelength)

        power_limit = settings["power_limit"]
        if not isinstance(power_limit, (int, float)) or power_limit <= 0:
            print(f"Error: 'power_limit' must be a positive number, got '{power_limit}'.")
            sys.exit(1)
        self.power_limit = float(power_limit)

        output_path = settings["output_path"]
        if not isinstance(output_path, str):
            print(f"Error: 'output_path' must be a string, got '{type(output_path).__name__}'.")
            sys.exit(1)
        self.output_path = Path(output_path)

# ---------------------------------------------------------------------------
# State handlers
# ---------------------------------------------------------------------------

def state_init(ctx: Context) -> State:
    print("\n=== INIT ===")
    ctx.current_angle = 0.0
    
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Measurement started at")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=datetime.now().isoformat())
    ctx.ws_row += 1

    ctx.resolve_output_path() # Output file
    return State.CHECK_DEVICE_1


def state_check_device1(ctx: Context) -> State:
    print("\n=== CHECK_DEVICE_1 ===")
    try:
        ok = ctx.dev1.connect()
        if not ok:
            raise RuntimeError("Device1 returned False")
        return State.CHECK_DEVICE_2
    except Exception as exc:
        print(f"  ERROR connecting to Device1: {exc}")
        return State.ERROR


def state_check_device2(ctx: Context) -> State:
    print("\n=== CHECK_DEVICE_2 ===")
    try:
        ok = ctx.dev2.connect()
        if not ok:
            raise RuntimeError("Device2 returned False")
        return State.LASER_CHECK
    except Exception as exc:
        print(f"  ERROR connecting to Device2: {exc}")
        return State.ERROR


def state_laser_check(ctx: Context) -> State:
    print("\n=== LASER_CHECK ===")
    answer = input("  Is the laser powered with Pout < 250 mW? [y/n]: ").strip().lower()
    if answer in ("y", "yes"):
        return State.COARSE_SCAN
    return State.QUIT


def state_coarse_scan(ctx: Context) -> State:
    """
    Sweep 100 deg in equal steps.
    Step size is 1 deg -> 100 steps total.
    """
    print("\n=== COARSE_SCAN (100 deg) ===")
    total_degrees = 100.0
    step_deg      = 1.0
    n_steps       = int(total_degrees / step_deg)

    ctx.coarse_angles.clear()
    ctx.coarse_powers.clear()

    for i in range(n_steps):
        power = ctx.dev2.measure_power()
        ctx.coarse_angles.append(ctx.current_angle)
        ctx.coarse_powers.append(power)

        ctx.dev1.rotate_step(step_deg)
        ctx.current_angle += step_deg

    print(f"  Coarse scan complete. {n_steps} points collected.")
    return State.FIND_MAX


def state_find_max(ctx: Context) -> State:
    print("\n=== FIND_MAX ===")
    idx = int(np.argmax(ctx.coarse_powers))
    ctx.max_position = ctx.coarse_angles[idx]
    print(f"  Maximum power {ctx.coarse_powers[idx]:.2f} mW at {ctx.max_position:.2f} deg")
    print(f"  Moving to {ctx.max_position:.2f} deg ...")
    ctx.dev1.go_to_position(ctx.max_position)
    ctx.current_angle = ctx.max_position
    return State.FINE_SCAN


def state_fine_scan(ctx: Context) -> State:
    """
    450 deg sweep at 0.25 deg/s, 10 measurement points per second.
    step_deg = rotation_rate / sample_rate = 0.25 / 10 = 0.025 deg/step
    """
    print("\n=== FINE_SCAN (450 deg @ 0.25 deg/s, 10 pts/s) ===")
    total_degrees  = 450.0
    rotation_rate  = 0.25   # deg/s
    sample_rate    = 10.0   # pts/s
    step_deg       = rotation_rate / sample_rate   # 0.025 deg
    n_steps        = int(total_degrees / step_deg)
    step_delay     = 1.0 / sample_rate             # seconds between samples

    ctx.fine_angles.clear()
    ctx.fine_powers.clear()

    start = time.time()
    for i in range(n_steps):
        t0 = time.time()

        power = ctx.dev2.measure_power()
        ctx.fine_angles.append(ctx.current_angle)
        ctx.fine_powers.append(power)

        ctx.dev1.rotate_step(step_deg)
        ctx.current_angle += step_deg

        # pace the loop to ~sample_rate
        elapsed = time.time() - t0
        sleep_remaining = step_delay - elapsed
        if sleep_remaining > 0:
            time.sleep(sleep_remaining)

    duration = time.time() - start
    print(f"  Fine scan complete. {n_steps} points in {duration:.1f}s.")
    return State.PLOT_POWER_ANGLE


def state_plot_power_angle(ctx: Context) -> State:
    print("\n=== PLOT_POWER_ANGLE ===")
    angles = ctx.fine_angles
    powers = ctx.fine_powers

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(angles, powers, linewidth=1.0, color="steelblue")
    ax.set_xlabel("Angle (deg)")
    ax.set_ylabel("Power (mW)")
    ax.set_title("Power vs Angle - Fine scan (450 deg)")
    ax.grid(True, alpha=0.3)

    # mark maximum
    idx_max = int(np.argmax(powers))
    ax.axvline(angles[idx_max], color="orangered", linestyle="--", linewidth=1,
               label=f"Max {powers[idx_max]:.1f} mW @ {angles[idx_max]:.2f} deg")
    ax.legend(fontsize=9)

    plt.tight_layout()
    plt.savefig("power_vs_angle.png", dpi=150)
    plt.show()
    print("  Plot saved: power_vs_angle.png")

    # store max position from fine scan
    ctx.max_position = angles[idx_max]
    return State.GOTO_MAX


def state_goto_max(ctx: Context) -> State:
    print("\n=== GOTO_MAX ===")
    print(f"  Moving to maximum position: {ctx.max_position:.2f} deg")
    ctx.dev1.go_to_position(ctx.max_position)
    ctx.current_angle = ctx.max_position
    return State.PRECISION_SCAN


def state_precision_scan(ctx: Context) -> State:
    """
    90 deg sweep at 0.5 deg/s, 10 pts/s.
    step_deg = 0.5 / 10 = 0.05 deg/step
    """
    print("\n=== PRECISION_SCAN (90 deg @ 0.5 deg/s, 10 pts/s) ===")
    total_degrees = 90.0
    rotation_rate = 0.5     # deg/s
    sample_rate   = 10.0    # pts/s
    step_deg      = rotation_rate / sample_rate   # 0.05 deg
    n_steps       = int(total_degrees / step_deg)
    step_delay    = 1.0 / sample_rate

    ctx.prec_angles.clear()
    ctx.prec_powers.clear()

    for i in range(n_steps):
        t0 = time.time()

        power = ctx.dev2.measure_power()
        ctx.prec_angles.append(ctx.current_angle)
        ctx.prec_powers.append(power)

        ctx.dev1.rotate_step(step_deg)
        ctx.current_angle += step_deg

        elapsed = time.time() - t0
        sleep_remaining = step_delay - elapsed
        if sleep_remaining > 0:
            time.sleep(sleep_remaining)

    print(f"  Precision scan complete. {n_steps} points.")

    # find minimum inside this 90 deg window for dwell position
    idx_min = int(np.argmin(ctx.prec_powers))
    ctx.min_position = ctx.prec_angles[idx_min]
    print(f"  Minimum power {ctx.prec_powers[idx_min]:.2f} mW at {ctx.min_position:.2f} deg")
    print(f"  Moving to minimum position for dwell ...")
    ctx.dev1.go_to_position(ctx.min_position)
    ctx.current_angle = ctx.min_position

    return State.DWELL


def state_dwell(ctx: Context) -> State:
    """
    Stay at minimum position for 30 minutes, 10 pts/s.
    """
    print("\n=== DWELL (30 min @ 10 pts/s) ===")
    dwell_duration = 30 * 60    # seconds
    sample_rate    = 10.0       # pts/s
    sample_delay   = 1.0 / sample_rate
    n_samples      = int(dwell_duration * sample_rate)

    ctx.dwell_times.clear()
    ctx.dwell_powers.clear()

    t_start = time.time()
    print(f"  Collecting {n_samples} samples over {dwell_duration/60:.0f} min ...")

    for i in range(n_samples):
        t0    = time.time()
        t_rel = t0 - t_start

        power = ctx.dev2.measure_power()
        ctx.dwell_times.append(t_rel)
        ctx.dwell_powers.append(power)

        if i % int(sample_rate * 60) == 0:
            mins = t_rel / 60
            print(f"  {mins:.1f} min elapsed - {power:.2f} mW")

        elapsed = time.time() - t0
        sleep_remaining = sample_delay - elapsed
        if sleep_remaining > 0:
            time.sleep(sleep_remaining)

    print("  Dwell complete.")
    return State.PLOT_POWER_TIME


def state_plot_power_time(ctx: Context) -> State:
    print("\n=== PLOT_POWER_TIME ===")
    times  = [t / 60.0 for t in ctx.dwell_times]   # convert to minutes
    powers = ctx.dwell_powers

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(times, powers, linewidth=0.8, color="darkorange")
    ax.set_xlabel("Time (min)")
    ax.set_ylabel("Power (mW)")
    ax.set_title(f"Power vs Time - 30-min dwell at {ctx.min_position:.2f} deg")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig("power_vs_time.png", dpi=150)
    plt.show()
    print("  Plot saved: power_vs_time.png")
    return State.OUTPUT_FILE


def state_output_file(ctx: Context) -> State:
    print(f"\n=== OUTPUT_FILE -> {ctx.output_path} ===")

    ctx.ws.cell(row=ctx.ws_row, column=1, value="Measurement finished at")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=datetime.now().isoformat())

    ctx.wb.save(ctx.output_path)

    print(f"  Excel file written: {ctx.output_path}")
    return State.DONE


# ---------------------------------------------------------------------------
# State machine runner
# ---------------------------------------------------------------------------

# Dispatch table
_HANDLERS = {
    State.INIT:             state_init,
    State.CHECK_DEVICE_1:   state_check_device1,
    State.CHECK_DEVICE_2:   state_check_device2,
    State.LASER_CHECK:      state_laser_check,
    State.COARSE_SCAN:      state_coarse_scan,
    State.FIND_MAX:         state_find_max,
    State.FINE_SCAN:        state_fine_scan,
    State.PLOT_POWER_ANGLE: state_plot_power_angle,
    State.GOTO_MAX:         state_goto_max,
    State.PRECISION_SCAN:   state_precision_scan,
    State.DWELL:            state_dwell,
    State.PLOT_POWER_TIME:  state_plot_power_time,
    State.OUTPUT_FILE:      state_output_file,
}

_TERMINAL_STATES = {State.DONE, State.QUIT, State.ERROR}


def run() -> None:
    parser = argparse.ArgumentParser(description="A console app that uses a settings file.") # 1. Set up the argument parser
    parser.add_argument( # 2. Add an argument for the config file (required)
        '-c', '--config',
        type=str,
        required=True,
        help="Path to the settings/configuration file."
    )
    args = parser.parse_args() # 3. Parse the arguments passed in the console

    print(f"Loading configuration from: {args.config}...\n")
    ctx = Context()
    ctx.load_settings

    state = State.INIT

    while state not in _TERMINAL_STATES:
        handler = _HANDLERS.get(state)
        if handler is None:
            print(f"ERROR: no handler for state {state}")
            state = State.ERROR
            break
        try:
            state = handler(ctx)
        except KeyboardInterrupt:
            print("\nInterrupted by user.")
            state = State.QUIT
            break
        except Exception as exc:
            print(f"\nUnhandled exception in state {state.name}: {exc}")
            state = State.ERROR
            break

    print(f"\n--- Final state: {state.name} ---")
    if state == State.QUIT:
        print("Measurement aborted by user.")
    elif state == State.ERROR:
        print("Measurement terminated due to an error.")
    elif state == State.DONE:
        print(f"Measurement complete. Data saved to: {ctx.output_path}")


if __name__ == "__main__":
    run()