"""
Laser Measurement State Machine
================================
States:
  INIT→ CHECK_POWERMETER → CHECK_LASER → CHEKC_ATT → WARMUP_LASER
  → LOOP(LOOP(RAMPUP_LASER → WAIT → MEASURE_POWER) → INCREASE_ATT)
  → OUTPUT_FILE → DONE
  (LASER_CHECK → QUIT on 'no')

"""

from __future__ import annotations

from openpyxl import Workbook # For Excel files
import time
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Optional
from typing import Any
import argparse
import json
import sys
import time
import ipaddress

from pymeasure.instruments.thorlabs import ThorlabsPM100USB
from Skyline.skyline import SkylineDevice, InterfaceType
from Thorlabs.FindThorlabs import find_thorlabs_pm100
from DA100.da100 import DA100Device

# ---------------------------------------------------------------------------
# State enum
# ---------------------------------------------------------------------------

class State(Enum):
    INIT            = auto()
    CHECK_POWERMETER= auto()
    CHECK_LASER     = auto()
    CHECK_ATT       = auto()
    WARMUP_LASER    = auto()
    RAMPUP_LASER    = auto()
    WAIT            = auto()
    MEASURE_POWER   = auto()
    INCREASE_ATT    = auto()
    OUTPUT_FILE     = auto()
    CLOSE_CONNECTIONS = auto()
    DONE            = auto()
    QUIT            = auto()
    ERROR           = auto()


# ---------------------------------------------------------------------------
# Measurement context  (shared data between states)
# ---------------------------------------------------------------------------

@dataclass
class Context:
    pm: Optional[ThorlabsPM100USB] = None
    skyline: Optional[SkylineDevice]    = None

    pm_sn:              Optional[str] = ""
    pm_wl:              float = 1500 # Unit [nm]

    skyline_cp:         str = ""
    P3_warmup_current:  Optional[float] = 0.0
    P3_max_current:     float = 0.0
    P3_step_current:    float = 0.0
    P3_min_current:     float = 0.0
    P3_current:         float = 0.0 # internal use only
    warmup_time:        int = 10 # Unit [s]
    settling_time:      int = 10 # Unit [s]

    att: Optional[DA100Device] = None
    att_cp:             str = ""
    att_min:            float = 0.0
    att_max:            float = 10.0
    att_step:           float = 2.0
    att_att:            float = 0.0 # internal use only

    step:               int = 0 # internal use only

    output_path:        Optional[Path] = None

    wb:                 Workbook = field(default_factory=Workbook)
    ws:                 Any = field(init=False)
    ws_row:             int = 1 # next row available (internal use only)

    def __post_init__(self):
        self.ws = self.wb.active

    def resolve_output_path(self) -> Path:
        if self.output_path is not None:
            return self.output_path
        
        i = 0
        while True:
            candidate = Path(f"alm_power_att{i:04d}.xlsx")
            if not candidate.exists():
                return candidate
            i += 1
    
    def load_settings(self, file_path: Path) -> None:
        """Reads, parses, and validates the JSON settings file into class members."""
        file_path = Path(file_path)

        if not file_path.exists():
            print(f"Error: Settings file '{file_path}' was not found.")
            sys.exit(1)

        if not file_path.suffix == ".json":
            print(f"Error: File '{file_path}' is not a JSON file.")
            sys.exit(1)

        try:
            with open(file_path, 'r') as f:
                settings = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Error: '{file_path}' is not valid JSON. Details: {e.msg} at line {e.lineno}.")
            sys.exit(1)

        self._validate_and_assign(settings)

    def _validate_and_assign_powermeter(self, settings: dict) -> None:
        required = ["wl"]
        for key in required:
            if key not in settings:
                print(f"Error: Missing required setting pm:'{key}'.")
                sys.exit(1)

        wavelength = settings["wl"]
        if not isinstance(wavelength, (int, float)) or not (200 <= wavelength <= 2000):
            print(f"Error: 'pm:wl' must be a number between 200 and 2000, got '{wavelength}'.")
            sys.exit(1)
        self.pm_wl = float(wavelength)

        self.pm_sn = str(settings["sn"]) if "sn" in settings else None
    
    def _validate_and_assign_skyline(self, settings: dict) -> None:
        required = ["cp", "P3_max_current", "P3_min_current", "P3_step_current", "warmup_time", "settling_time"]
        for key in required:
            if key not in settings:
                print(f"Error: Missing required setting skyline:'{key}'.")
                sys.exit(1)
        
        comport = settings["cp"]
        if not isinstance(comport, str):
            print(f"Error: 'skyline:cp' must be a string, got '{comport}'.")
            sys.exit(1)
        if len(comport) < 4 or len(comport) > 5:
            print(f"Error: 'skyline:cp' must provide only the COM port (like: 'COM5'), got {comport}")
            sys.exit(1)
        self.skyline_cp = str(comport)

        max_current = settings["P3_max_current"]
        if not isinstance(max_current, (int, float)) or max_current <= 0:
            print(f"Error: 'skyline:P3_max_current' must be a positive number, got '{max_current}'.")
            sys.exit(1)
        self.P3_max_current = float(max_current)

        min_current = settings["P3_min_current"]
        if not isinstance(min_current, (int, float)) or min_current <= 0:
            print(f"Error: 'skyline:P3_min_current' must be a positive number, got '{min_current}'.")
            sys.exit(1)
        if min_current >= max_current:
            print(f"Error: 'skyline:P3_min_current' supposed to be greater than 'skyline:P3_max_current'")
            sys.exit(1)
        self.P3_min_current = float(min_current)

        if "P3_warmup_current" in settings:
            warmup_current = settings["P3_warmup_current"]
            if not isinstance(warmup_current, (int, float)) or warmup_current < 0:
                print(f"Error: 'skyline:P3_warmup_current' should be a non-negative number, got {warmup_current}")
                sys.exit(1)
            if warmup_current > self.P3_min_current:
                print("Warning: skyline:P3_warmup_current is higher than skyline:P3_min_current")
            self.P3_warmup_current = float(warmup_current)

        current_step = settings["P3_step_current"]
        if not isinstance(current_step, (int, float)) or current_step <= 0:
            print(f"Error: 'skyline:P3_step_current' should be greater than 0, got {current_step}")
            sys.exit(1)
        self.P3_step_current = current_step

        warmup_time = settings["warmup_time"]
        if not isinstance(warmup_time, int) or warmup_time < 0:
            print(f"Error: 'skyline:warmup_time' should be greater than 0, got {warmup_time}")
            sys.exit(1)
        self.warmup_time = warmup_time

        settling_time = settings["settling_time"]
        if not isinstance(settling_time, int) or settling_time < 0:
            print(f"Error: 'skyline:settling_time' should be greater than 0, got {settling_time}")
            sys.exit(1)
        self.settling_time = settling_time

    def _validate_and_assign_attenuator(self, settings: dict) -> None:
        required = ["cp", "min", "max", "step"]
        for key in required:
            if key not in settings:
                print(f"Error: Missing required setting att:'{key}'.")
                sys.exit(1)
        
        comport = settings["cp"]
        if not isinstance(comport, str):
            print(f"Error: 'att:cp' must be a string, got '{comport}'.")
            sys.exit(1)
        if len(comport) < 4 or len(comport) > 5:
            print(f"Error: 'att:cp' must provide only the COM port (like: 'COM5'), got {comport}")
            sys.exit(1)
        self.att_cp = str(comport)

        min = settings["min"]
        if not isinstance(min, (int, float)) or min < 0:
            print(f"Error: att:min value not valid, got '{min}'")
            sys.exit(1)
        self.att_min = min

        max = settings["max"]
        if not isinstance(max, (int, float)) or max <= 0 or max < min:
            print(f"Error: att:max value not valid, got '{max}'")
            sys.exit(1)
        self.att_max = max

        step = settings["step"]
        if not isinstance(step, (int, float)) or step <= 0 or step > max-min:
            print(f"Error: att:step value not valid, got '{step}'")
            sys.exit(1)
        self.att_step = step

    def _validate_and_assign(self, settings: dict) -> None:
        """Validates and assigns settings to class members."""

        # --- Required fields ---
        required = ["pm", "skyline", "att"]
        for key in required:
            if key not in settings:
                print(f"Error: Missing required setting '{key}'.")
                sys.exit(1)

        # Power meter Thorlabs PM100 input parameters
        self._validate_and_assign_powermeter(settings["pm"])

        # Skyline input parameters
        self._validate_and_assign_skyline(settings["skyline"])
        
        # Attenuator input parameters
        self._validate_and_assign_attenuator(settings["att"])

        if "output_path" in settings:
            output_path = settings["output_path"]
            if not isinstance(output_path, str):
                print(f"Error: 'output_path' must be a string, got '{type(output_path).__name__}'.")
                sys.exit(1)
            self.output_path = Path(output_path)

# ---------------------------------------------------------------------------
# State handlers
# ---------------------------------------------------------------------------

def state_init(ctx: Context) -> State:
    print("\n=== INIT ===")
    
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Measurement started at")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=datetime.now().isoformat())
    ctx.ws_row += 1

    ctx.output_path = ctx.resolve_output_path() # Output file
    return State.CHECK_POWERMETER
    
def state_check_powermeter(ctx: Context) -> State:
    print("=== CHECK_POWERMETER ===")
    
    devices = find_thorlabs_pm100()
    if len(devices) == 0:
        print("No Thorlabs PM100 found.")
        sys.exit(1)
    
    if ctx.pm_sn is None:
        if  len(devices) == 1:
            d = devices[0]
            print(f"  No serial number specified, using first found: {d['model']} SN:{d['serial']}")
            ctx.pm = ThorlabsPM100USB(d["resource_string"])
            ctx.pm.wavelength = ctx.pm_wl
        else:
            print(f"Error: no power meter serial number specified and more than one connected")
    else:
        matched = next((d for d in devices if d['serial'] == ctx.pm_sn), None)
        if matched is None:
            print(f"Error: No PM100 with serial '{ctx.pm_sn}' found.")
            print(f"  Available devices: {[d['serial'] for d in devices]}")
            sys.exit(1)
        ctx.pm = ThorlabsPM100USB(matched["resource_string"])
        ctx.pm.wavelength = ctx.pm_wl

    assert ctx.pm is not None, "Power meter not initialized"
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Power meter info")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=f"{ctx.pm.id},{ctx.pm.name},{ctx.pm.sensor_name},{ctx.pm.sensor_sn},{ctx.pm.wavelength}")
    ctx.ws_row += 1

    return State.CHECK_LASER

def state_check_laser(ctx: Context) -> State:
    print("\n=== CHECK_LASER ===")
    
    address = f"{ctx.skyline_cp}:57600:8:N:1"
    print(f"  Connecting to Skyline at: {address}")

    ctx.skyline = SkylineDevice(InterfaceType.RS232, address)
    ctx.skyline.__enter__()
    info = ctx.skyline.get_device_info()

    ctx.ws.cell(row=ctx.ws_row, column=1, value="Skyline")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=f"{info['serial']},{info['firmware']},{ctx.skyline.get_cpld_version()}")
    ctx.ws_row += 1

    return State.CHECK_ATT

def state_check_attenuator(ctx: Context) -> State:
    print("=== CHECK_ATT ===")

    address = f"{ctx.att_cp}:9600:8:N:1"
    #print(f"  Connecting to the attenuator at: '{address}'")
    ctx.att = DA100Device(address)
    ctx.att.__enter__()
    info = ctx.att.get_device_name_and_version()
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Attenuator info")
    ctx.ws_row += 1
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Device name")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=info['name'])
    ctx.ws_row += 1
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Version")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=info['sw_version'])
    ctx.ws.cell(row=ctx.ws_row, column=3, value=info['hw_version'])
    ctx.ws_row += 1
    ctx.ws.cell(row=ctx.ws_row, column=1, value="SN")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=ctx.att.get_serial_number())
    ctx.ws_row += 2

    ctx.att_att = ctx.att_min
    print(f"Attenuation set to: {ctx.att_att}")
    ctx.att.set_attenuation(ctx.att_att)
    ctx.ws.cell(row=ctx.ws_row, column=1, value="Attenuation")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=ctx.att_att)
    ctx.ws_row += 1

    return State.WARMUP_LASER

def state_warmup_laser(ctx: Context) -> State:
    assert ctx.skyline is not None, "Skyline device not initialized"
    ctx.skyline.set_ao_pump3_current(0.0)
    ctx.skyline.set_pump3_enabled(True)
    if isinstance(ctx.P3_warmup_current, float):
        ctx.skyline.set_ao_pump3_current(ctx.P3_warmup_current)
        ctx.P3_current = ctx.P3_warmup_current
    else:
        ctx.P3_current = 0.0
    print(f"=== WARMUP_LASER ({ctx.warmup_time} s) ===")
    time.sleep(ctx.warmup_time) # Time in seconds

    return State.RAMPUP_LASER

def state_rampup_laser(ctx: Context) -> State:
    assert ctx.skyline is not None, "Skyline not initialized"
    current = ctx.P3_min_current + ctx.step * ctx.P3_step_current
    print(f"=== RAMPUP_LASER (I={current} A) ===")
    ctx.skyline.set_ao_pump3_current(current)
    ctx.P3_current = current
    ctx.step += 1

    return State.WAIT

def state_wait(ctx: Context) -> State:
    print(f"=== WAIT ({ctx.settling_time} s) ===")
    
    time.sleep(ctx.settling_time) # Time in seconds

    return State.MEASURE_POWER

def state_measure_power(ctx: Context) -> State:
    assert ctx.pm is not None, "Power meter not initialized"
    
    power = ctx.pm.power    
    print(f"=== MEASURE_POWER (I={ctx.P3_current}A, P={power}) ===")

    ctx.ws.cell(row=ctx.ws_row, column=1, value=ctx.P3_current)
    ctx.ws.cell(row=ctx.ws_row, column=2, value=power)
    ctx.ws_row += 1

    if ctx.P3_current + ctx.P3_step_current > ctx.P3_max_current:
        ctx.skyline.set_ao_pump3_current(ctx.P3_min_current)
        return State.INCREASE_ATT
    else:
        return State.RAMPUP_LASER
    
def state_increase_attenuation(ctx: Context) -> State:
    assert ctx.att is not None, "Attenuator not initialized"

    if ctx.att_att + ctx.att_step > ctx.att_max:
        return State.OUTPUT_FILE
    else:
        ctx.ws_row += 1
        ctx.att_att = ctx.att_att + ctx.att_step
        print(f"=== INCREASE_ATT ({ctx.att_att} dB) ===")
        ctx.att.set_attenuation(ctx.att_att)
        ctx.ws.cell(row=ctx.ws_row, column=1, value="Attenuation")
        ctx.ws.cell(row=ctx.ws_row, column=2, value=ctx.att_att)
        ctx.ws_row += 1

        ctx.step = 0
    
        return State.WARMUP_LASER

def state_output_file(ctx: Context) -> State:
    print(f"=== OUTPUT_FILE -> {ctx.output_path} ===")

    ctx.ws.cell(row=ctx.ws_row, column=1, value="Measurement finished at")
    ctx.ws.cell(row=ctx.ws_row, column=2, value=datetime.now().isoformat())

    assert ctx.output_path is not None, "Output path not properly set"
    ctx.wb.save(ctx.output_path)

    print(f"  Excel file written: {ctx.output_path}")
    return State.CLOSE_CONNECTIONS

def state_close_connections(ctx: Context) -> State:
    print("=== CLOSE_CONNECTIONS ===")

    assert ctx.pm is not None, "Power meter not initialized"
    assert ctx.skyline is not None, "Skyline not initialized"

    try:
        ctx.skyline.set_ao_pump3_current(0.0)
        ctx.skyline.set_pump3_enabled(False)
    except Exception as e:
        print(f"  Failed to close Skyline connection: {e}")

    try:
        ctx.pm.shutdown()
    except Exception as e:
        print(f"  Failed to close power meter connection: {e}")

    return State.DONE


# ---------------------------------------------------------------------------
# State machine runner
# ---------------------------------------------------------------------------

# Dispatch table
_HANDLERS = {
    State.INIT:             state_init,
    State.CHECK_POWERMETER: state_check_powermeter,
    State.CHECK_LASER:      state_check_laser,
    State.CHECK_ATT:        state_check_attenuator,
    State.WARMUP_LASER:     state_warmup_laser,
    State.RAMPUP_LASER:     state_rampup_laser,
    State.WAIT:             state_wait,
    State.MEASURE_POWER:    state_measure_power,
    State.INCREASE_ATT:     state_increase_attenuation,
    State.OUTPUT_FILE:      state_output_file,
    State.CLOSE_CONNECTIONS:state_close_connections,
}

_TERMINAL_STATES = {State.DONE, State.QUIT, State.ERROR}


def run() -> None:
    parser = argparse.ArgumentParser(description="A console app that uses a settings file.") # 1. Set up the argument parser
    parser.add_argument( # 2. Add an argument for the config file (required)
        '-c', '--config',
        type=str,
        required=True,
        help="Path to the settings/configuration file."
    )
    args = parser.parse_args() # 3. Parse the arguments passed in the console

    print(f"Loading configuration from: {args.config}...\n")
    ctx = Context()
    ctx.load_settings(args.config)

    state = State.INIT

    while state not in _TERMINAL_STATES:
        handler = _HANDLERS.get(state)
        if handler is None:
            print(f"ERROR: no handler for state {state}")
            state = State.ERROR
            break
        try:
            state = handler(ctx)
        except KeyboardInterrupt:
            print("\nInterrupted by user.")
            state = State.OUTPUT_FILE
            break
        except Exception as exc:
            print(f"\nUnhandled exception in state {state.name}: {exc}")
            state = State.ERROR
            break

    print(f"\n--- Final state: {state.name} ---")
    if state == State.QUIT:
        print("Measurement aborted by user.")
    elif state == State.ERROR:
        print("Measurement terminated due to an error.")
    elif state == State.DONE:
        print(f"Measurement complete. Data saved to: {ctx.output_path}")


if __name__ == "__main__":
    run()